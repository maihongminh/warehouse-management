from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models import Batch, Product
from app.schemas.batch import BatchOut
from app.schemas.product import ProductCreate, ProductOut, ProductUpdate

router = APIRouter()


@router.get("", response_model=list[ProductOut])
def list_products(
    db: Session = Depends(get_db),
    active_only: bool = Query(True),
    q: str | None = Query(None, description="Search name or sku"),
) -> list[Product]:
    stmt = db.query(Product)
    if active_only:
        stmt = stmt.filter(Product.is_active.is_(True))
    if q:
        like = f"%{q}%"
        stmt = stmt.filter((Product.name.ilike(like)) | (Product.sku.ilike(like)))
    return stmt.order_by(Product.name.asc()).all()


@router.post("", response_model=ProductOut, status_code=201)
def create_product(body: ProductCreate, db: Session = Depends(get_db)) -> Product:
    exists = db.query(Product).filter(Product.sku == body.sku).first()
    if exists:
        raise HTTPException(status_code=400, detail="Mã SKU đã tồn tại")
    p = Product(
        name=body.name,
        sku=body.sku,
        unit=body.unit,
        default_import_price=body.default_import_price,
        default_sale_price=body.default_sale_price,
        conversion_rate=body.conversion_rate,
        is_active=body.is_active,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return p


@router.get("/export-excel")
def export_products_excel(db: Session = Depends(get_db)):
    """Xuất danh sách sản phẩm ra file Excel theo chuẩn mẫu."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    products = db.query(Product).filter(Product.is_active.is_(True)).order_by(Product.name.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DanhSachSanPham"

    headers = ["Mã hàng", "Tên hàng", "Giá bán", "Giá vốn", "Tồn kho", "ĐVT", "Quy đổi", "Lô 1", "Hạn sử dụng 1"]
    header_fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(products, 2):
        batches = db.query(Batch).filter(
            Batch.product_id == p.id,
            Batch.quantity_remaining > 0
        ).order_by(Batch.expiry_date.asc()).all()
        total_qty = sum(b.quantity_remaining for b in batches)

        # Lô gần hết hạn nhất (FEFO) để xuất vào cột Lô 1 / HSD 1
        first_batch = batches[0] if batches else None
        batch_code = first_batch.batch_code or "" if first_batch else ""
        expiry_date = first_batch.expiry_date if first_batch else None

        ws.cell(row=row_idx, column=1, value=p.sku)
        ws.cell(row=row_idx, column=2, value=p.name)
        ws.cell(row=row_idx, column=3, value=float(p.default_sale_price))
        ws.cell(row=row_idx, column=4, value=float(p.default_import_price))
        ws.cell(row=row_idx, column=5, value=total_qty)
        ws.cell(row=row_idx, column=6, value=p.unit)
        ws.cell(row=row_idx, column=7, value=p.conversion_rate)
        ws.cell(row=row_idx, column=8, value=batch_code)
        # Ghi ngày dạng date object để Excel tự nhận biết định dạng ngày
        if expiry_date:
            from datetime import date
            cell_exp = ws.cell(row=row_idx, column=9, value=expiry_date)
            cell_exp.number_format = "DD/MM/YYYY"
        else:
            ws.cell(row=row_idx, column=9, value="")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=DanhSachSanPham.xlsx"},
    )



@router.post("/import-excel")
def import_products_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Nhập hàng loạt sản phẩm từ file Excel. Upsert theo SKU."""
    import openpyxl

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Chỉ chấp nhận file .xlsx")

    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không đọc được file Excel: {e}") from e

    created = 0
    updated = 0
    errors: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Cột: Mã hàng | Tên hàng | Giá bán | Giá vốn | Tồn kho | ĐVT | Quy đổi
        if not row or row[0] is None:
            continue
        try:
            sku = str(row[0]).strip()
            name = str(row[1]).strip() if row[1] else ""
            sale_price = float(row[2] or 0)
            import_price = float(row[3] or 0)
            unit = str(row[5]).strip() if row[5] else "unit"
            conversion_rate = int(row[6] or 1)

            if not sku or not name:
                errors.append(f"Dòng {row_idx}: SKU hoặc tên trống, bỏ qua.")
                continue

            existing = db.query(Product).filter(Product.sku == sku).first()
            if existing:
                existing.name = name
                existing.default_sale_price = sale_price
                existing.default_import_price = import_price
                existing.unit = unit
                existing.conversion_rate = conversion_rate
                existing.is_active = True
                updated += 1
            else:
                p = Product(
                    sku=sku,
                    name=name,
                    default_sale_price=sale_price,
                    default_import_price=import_price,
                    unit=unit,
                    conversion_rate=conversion_rate,
                    is_active=True,
                )
                db.add(p)
                created += 1
        except Exception as e:
            errors.append(f"Dòng {row_idx}: {e}")

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "errors": errors,
        "message": f"Tạo mới {created}, cập nhật {updated} sản phẩm. {len(errors)} lỗi.",
    }


@router.get("/{product_id}", response_model=ProductOut)
def get_product(product_id: int, db: Session = Depends(get_db)) -> Product:
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy sản phẩm")
    return p


@router.get("/{product_id}/batches", response_model=list[BatchOut])
def list_product_batches(product_id: int, db: Session = Depends(get_db)) -> list[Batch]:
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy sản phẩm")
    return (
        db.query(Batch)
        .filter(Batch.product_id == product_id, Batch.quantity_remaining > 0)
        .order_by(Batch.expiry_date.asc(), Batch.id.asc())
        .all()
    )


@router.patch("/{product_id}", response_model=ProductOut)
def update_product(product_id: int, body: ProductUpdate, db: Session = Depends(get_db)) -> Product:
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy sản phẩm")
    data = body.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(p, k, v)
    db.commit()
    db.refresh(p)
    return p


@router.delete("/{product_id}", status_code=204)
def delete_product(product_id: int, db: Session = Depends(get_db)) -> None:
    """Soft delete — đánh dấu is_active=False, không xóa dữ liệu thật."""
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy sản phẩm")
    p.is_active = False
    db.commit()
